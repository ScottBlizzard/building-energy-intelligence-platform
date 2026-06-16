# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[3]
XLSX_PATH = ROOT / "final_docs" / "图象重绘" / "源代码" / "SEM-6.6-甘特图.xlsx"
PNG_PATH = ROOT / "final_docs" / "images" / "SEM-6.6-甘特图.png"

WEEKS = [
    ("第1周", "04/19-04/25"),
    ("第2周", "04/26-05/02"),
    ("第3周", "05/03-05/09"),
    ("第4周", "05/10-05/16"),
    ("第5周", "05/17-05/23"),
    ("第6周", "05/24-05/30"),
    ("第7周", "05/31-06/06"),
    ("第8周", "06/07-06/13"),
]

TASKS = [
    {
        "wbs": "1.0",
        "name": "项目文档与范围基线",
        "plan": (1, 7),
        "actual": (1, 8),
        "note": "封版文档吸收闭环、ROI、MySQL 与数据口径变更",
    },
    {
        "wbs": "2.0",
        "name": "工程基础与运行环境",
        "plan": (1, 1),
        "actual": (1, 2),
        "note": "后端依赖和启动说明补齐",
    },
    {
        "wbs": "3.0",
        "name": "数据资产与知识库",
        "plan": (1, 3),
        "actual": (1, 5),
        "note": "L0-L3 数据质量、知识库和经济口径继续收敛",
    },
    {
        "wbs": "4.0",
        "name": "用户认证与角色权限",
        "plan": (8, 8),
        "actual": (7, 8),
        "note": "角色权限提前接入前端联调，按期封版",
    },
    {
        "wbs": "5.0",
        "name": "数据查询与统计分析",
        "plan": (1, 6),
        "actual": (1, 7),
        "note": "统计分析、图表和字段联调延伸一周",
    },
    {
        "wbs": "6.0",
        "name": "业务闭环与时间沙盘",
        "plan": (8, 8),
        "actual": (7, 8),
        "note": "状态机与沙盘提前预研，第 8 周集中验收",
    },
    {
        "wbs": "7.0",
        "name": "预算、KPI 与 ROI 决策",
        "plan": (8, 8),
        "actual": (7, 8),
        "note": "预算口径和 ROI 方法学提前联调，按期完成",
    },
    {
        "wbs": "8.0",
        "name": "AI 助手与 MCP 接入",
        "plan": (7, 7),
        "actual": (7, 8),
        "note": "MCP 基础后继续补接地问答、降级和质量门控",
    },
    {
        "wbs": "9.0",
        "name": "前端工作台与可视化",
        "plan": (1, 7),
        "actual": (1, 8),
        "note": "角色页、预算 ROI 展示和演示体验延伸一周",
    },
    {
        "wbs": "10.0",
        "name": "持久化与数据库适配",
        "plan": (8, 8),
        "actual": (7, 8),
        "note": "仓储适配提前预研，第 8 周完成 MySQL 可选增强",
    },
    {
        "wbs": "11.0",
        "name": "测试、验收与结项交付",
        "plan": (7, 8),
        "actual": (6, 8),
        "note": "测试前移，封版周完成验收与提交材料",
    },
]

COLORS = {
    "title": "17324D",
    "header": "EAF2F8",
    "plan": "4F7FD1",
    "early": "70AD7D",
    "actual": "F3A047",
    "late": "D9534F",
    "grid": "D9E2EC",
    "label": "F7FAFC",
}


def fill(color):
    return PatternFill("solid", fgColor=color)


def generate_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "SEM-6.6-Gantt"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    ws["A1"] = "建筑能源智能管理系统一级 WBS 计划-实际甘特图"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color=COLORS["title"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["WBS", "工作包", "轨迹"] + [f"{week}\n{date}" for week, date in WEEKS]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="1F3347")
        cell.fill = fill(COLORS["header"])
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    start_row = 4
    for idx, task in enumerate(TASKS):
        plan_row = start_row + idx * 2
        actual_row = plan_row + 1
        ws.merge_cells(start_row=plan_row, start_column=1, end_row=actual_row, end_column=1)
        ws.merge_cells(start_row=plan_row, start_column=2, end_row=actual_row, end_column=2)
        ws.cell(plan_row, 1, task["wbs"])
        ws.cell(plan_row, 2, task["name"])
        ws.cell(plan_row, 3, "计划")
        ws.cell(actual_row, 3, "实际")

        for row in (plan_row, actual_row):
            ws.row_dimensions[row].height = 22
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(name="微软雅黑", size=9, color="24364A")
                if col <= 3:
                    cell.fill = fill(COLORS["label"])

        plan_start, plan_end = task["plan"]
        actual_start, actual_end = task["actual"]
        for week in range(1, 9):
            col = 3 + week
            plan_cell = ws.cell(plan_row, col)
            actual_cell = ws.cell(actual_row, col)
            if plan_start <= week <= plan_end:
                plan_cell.fill = fill(COLORS["plan"])
                plan_cell.value = "计划"
                plan_cell.font = Font(name="微软雅黑", size=8, color="FFFFFF", bold=True)
            if actual_start <= week <= actual_end:
                if week < plan_start:
                    color, label = COLORS["early"], "提前"
                elif week > plan_end:
                    color, label = COLORS["late"], "延期"
                else:
                    color, label = COLORS["actual"], "实际"
                actual_cell.fill = fill(color)
                actual_cell.value = label
                actual_cell.font = Font(name="微软雅黑", size=8, color="FFFFFF", bold=True)

    legend_row = start_row + len(TASKS) * 2 + 2
    ws.cell(legend_row, 1, "图例").font = Font(name="微软雅黑", size=10, bold=True)
    legend_items = [
        ("计划", COLORS["plan"]),
        ("实际提前开始", COLORS["early"]),
        ("计划窗口内实际执行", COLORS["actual"]),
        ("实际晚于计划", COLORS["late"]),
    ]
    for idx, (label, color) in enumerate(legend_items):
        col = 2 + idx * 2
        ws.cell(legend_row, col, "").fill = fill(color)
        ws.cell(legend_row, col).border = border
        ws.cell(legend_row, col + 1, label).font = Font(name="微软雅黑", size=9, color="24364A")
        ws.cell(legend_row, col + 1).alignment = Alignment(vertical="center")

    note_row = legend_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=11)
    ws.cell(
        note_row,
        1,
        "说明：本图按一级 WBS 展示计划与实际双轨。绿色表示实际工作早于计划窗口启动，红色表示实际执行超过计划结束周；实际早于计划不改变基线，只用于说明团队提前预研或测试前移。",
    )
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(note_row, 1).font = Font(name="微软雅黑", size=9, color="53687A")

    widths = [8, 30, 8] + [14] * 8
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "D4"
    ws.print_area = f"A1:K{note_row + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    wb.save(XLSX_PATH)


def text_center(draw, xy, text, font, fill_color="#24364A"):
    x0, y0, x1, y1 = xy
    box = draw.textbbox((0, 0), text, font=font)
    text_width, text_height = box[2] - box[0], box[3] - box[1]
    draw.text((x0 + (x1 - x0 - text_width) / 2, y0 + (y1 - y0 - text_height) / 2 - 1), text, font=font, fill=fill_color)


def generate_png():
    font_regular = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 22)
    font_small = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 18)
    font_tiny = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 15)
    font_bold = ImageFont.truetype("C:/Windows/Fonts/msyhbd.ttc", 28)
    font_header = ImageFont.truetype("C:/Windows/Fonts/msyhbd.ttc", 18)

    width, height = 1600, 900
    image = Image.new("RGB", (width, height), "#F8FBFD")
    draw = ImageDraw.Draw(image)

    margin_x = 45
    top = 105
    row_h = 28
    head_h = 58
    col_w = {"wbs": 80, "name": 365, "type": 70, "week": 125}
    x_wbs = margin_x
    x_name = x_wbs + col_w["wbs"]
    x_type = x_name + col_w["name"]
    x_week0 = x_type + col_w["type"]
    chart_right = x_week0 + col_w["week"] * 8

    draw.text((margin_x, 34), "建筑能源智能管理系统一级 WBS 计划-实际甘特图", font=font_bold, fill="#17324D")
    draw.text((margin_x, 72), "蓝色为计划基线，绿色为提前启动，橙色为计划窗口内实际执行，红色为晚于计划的延期周。", font=font_small, fill="#53687A")

    draw.rounded_rectangle((margin_x, top, chart_right, top + head_h), radius=8, fill="#EAF2F8", outline="#C8D6E2")
    for x in [x_name, x_type, x_week0] + [x_week0 + idx * col_w["week"] for idx in range(1, 8)]:
        draw.line((x, top, x, top + head_h + len(TASKS) * 2 * row_h), fill="#D6E1EA", width=1)
    text_center(draw, (x_wbs, top, x_name, top + head_h), "WBS", font_header)
    text_center(draw, (x_name, top, x_type, top + head_h), "工作包", font_header)
    text_center(draw, (x_type, top, x_week0, top + head_h), "轨迹", font_header)
    for idx, (week, date) in enumerate(WEEKS):
        x0 = x_week0 + idx * col_w["week"]
        x1 = x0 + col_w["week"]
        text_center(draw, (x0, top + 6, x1, top + 28), week, font_header)
        text_center(draw, (x0, top + 30, x1, top + head_h), date, font_tiny, "#53687A")

    bar_colors = {"plan": "#4F7FD1", "early": "#70AD7D", "actual": "#F3A047", "late": "#D9534F"}
    row_top = top + head_h
    for idx, task in enumerate(TASKS):
        y_plan = row_top + idx * 2 * row_h
        y_actual = y_plan + row_h
        draw.rectangle((margin_x, y_plan, chart_right, y_actual + row_h), fill="#FFFFFF", outline="#D6E1EA")
        draw.rectangle((margin_x, y_plan, x_week0, y_actual + row_h), fill="#F7FAFC", outline="#D6E1EA")
        draw.line((margin_x, y_actual, chart_right, y_actual), fill="#E9EEF3", width=1)
        draw.line((margin_x, y_actual + row_h, chart_right, y_actual + row_h), fill="#D6E1EA", width=1)
        text_center(draw, (x_wbs, y_plan, x_name, y_actual + row_h), task["wbs"], font_small)
        draw.text((x_name + 10, y_plan + 17), task["name"], font=font_small, fill="#24364A")
        text_center(draw, (x_type, y_plan, x_week0, y_actual), "计划", font_tiny)
        text_center(draw, (x_type, y_actual, x_week0, y_actual + row_h), "实际", font_tiny)

        plan_start, plan_end = task["plan"]
        actual_start, actual_end = task["actual"]
        for week in range(1, 9):
            x0 = x_week0 + (week - 1) * col_w["week"] + 7
            x1 = x0 + col_w["week"] - 14
            if plan_start <= week <= plan_end:
                draw.rounded_rectangle((x0, y_plan + 5, x1, y_plan + row_h - 5), radius=4, fill=bar_colors["plan"])
            if actual_start <= week <= actual_end:
                if week < plan_start:
                    color, label = bar_colors["early"], "提前"
                elif week > plan_end:
                    color, label = bar_colors["late"], "延期"
                else:
                    color, label = bar_colors["actual"], "实际"
                draw.rounded_rectangle((x0, y_actual + 5, x1, y_actual + row_h - 5), radius=4, fill=color)
                text_center(draw, (x0, y_actual + 5, x1, y_actual + row_h - 5), label, font_tiny, "#FFFFFF")

    legend_y = row_top + len(TASKS) * 2 * row_h + 24
    legend_x = margin_x
    for label, color in [
        ("计划", bar_colors["plan"]),
        ("实际提前开始", bar_colors["early"]),
        ("计划窗口内实际执行", bar_colors["actual"]),
        ("实际晚于计划", bar_colors["late"]),
    ]:
        draw.rounded_rectangle((legend_x, legend_y, legend_x + 34, legend_y + 18), radius=4, fill=color)
        draw.text((legend_x + 44, legend_y - 4), label, font=font_tiny, fill="#24364A")
        legend_x += 210

    note = "复盘重点：WBS 3.0、5.0、8.0、9.0 出现延期；WBS 4.0、6.0、7.0、10.0、11.0 通过提前预研或测试前移抵消了第 8 周集中封版压力。"
    draw.rounded_rectangle((margin_x, legend_y + 38, chart_right, legend_y + 88), radius=8, fill="#FFFFFF", outline="#D6E1EA")
    draw.text((margin_x + 18, legend_y + 52), note, font=font_small, fill="#53687A")
    image.save(PNG_PATH)


if __name__ == "__main__":
    generate_workbook()
    generate_png()
    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {PNG_PATH}")
